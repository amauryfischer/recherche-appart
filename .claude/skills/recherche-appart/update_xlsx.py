#!/usr/bin/env python3
"""Met à jour annonces.xlsx à partir d'annonces au format JSON sur stdin.

Usage:
    update_xlsx.py --list          # liste les liens déjà connus (un par ligne)
    echo '[...]' | update_xlsx.py  # fusionne les annonces fournies
"""

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX = Path(__file__).resolve().parents[3] / "annonces.xlsx"

COLUMNS = [
    ("Date ajout", 12),
    ("Nouveau", 9),
    ("Prix", 12),
    ("Surface m²", 11),
    ("Prix/m²", 10),
    ("Pièces", 8),
    ("Étage", 10),
    ("Ascenseur", 11),
    ("DPE", 6),
    ("Charges", 10),
    ("Rue / quartier", 24),
    ("Description", 50),
    ("Source", 16),
    ("Lien", 45),
    ("Notes", 30),
]

LINK_COL = [c[0] for c in COLUMNS].index("Lien") + 1
NEW_COL = [c[0] for c in COLUMNS].index("Nouveau") + 1


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Annonces"
    ws.append([name for name, _ in COLUMNS])
    header_fill = PatternFill("solid", fgColor="1F3864")
    for idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    return wb


def load():
    if XLSX.exists():
        return load_workbook(XLSX)
    return create_workbook()


def known_links(ws):
    return {
        ws.cell(row=r, column=LINK_COL).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=LINK_COL).value
    }


def price_per_m2(prix, surface):
    try:
        return round(float(prix) / float(surface))
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def main():
    wb = load()
    ws = wb.active

    if "--list" in sys.argv:
        for link in sorted(known_links(ws)):
            print(link)
        return

    raw = sys.stdin.read().strip()
    if not raw:
        print("Aucune annonce reçue sur stdin.", file=sys.stderr)
        sys.exit(1)

    annonces = json.loads(raw)
    if isinstance(annonces, dict):
        annonces = [annonces]

    seen = known_links(ws)

    # Les trouvailles du run précédent ne sont plus "nouvelles".
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=NEW_COL).value = None

    today = date.today().isoformat()
    added = 0

    for a in annonces:
        lien = (a.get("lien") or "").strip()
        if not lien or lien in seen:
            continue
        seen.add(lien)
        added += 1
        ws.append([
            today,
            "★",
            a.get("prix"),
            a.get("surface_m2"),
            price_per_m2(a.get("prix"), a.get("surface_m2")),
            a.get("pieces"),
            a.get("etage"),
            a.get("ascenseur"),
            a.get("dpe"),
            a.get("charges_mensuelles"),
            a.get("rue_ou_quartier"),
            a.get("description"),
            a.get("source"),
            lien,
            None,  # Notes — réservé à l'utilisateur
        ])
        row = ws.max_row
        ws.cell(row=row, column=NEW_COL).font = Font(bold=True, color="C00000")
        ws.cell(row=row, column=12).alignment = Alignment(wrap_text=True, vertical="top")

    for col in (3, 5, 10):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "#,##0 €"

    wb.save(XLSX)
    print(f"{added} nouvelle(s) annonce(s) — {ws.max_row - 1} au total dans {XLSX.name}")


if __name__ == "__main__":
    main()
